# -*- coding: utf-8 -*-
import os
import shutil
from asyncio import sleep
import datetime

import pandas as pd
################################################################################
## Form generated from reading UI file 'spravochnik_papers.ui'
##
## Created by: Qt User Interface Compiler version 6.7.2
##
## WARNING! All changes made in this file will be lost when recompiling UI file!
################################################################################

from PySide6.QtCore import (QCoreApplication, QDate, QDateTime, QLocale,
    QMetaObject, QObject, QPoint, QRect,
    QSize, QTime, QUrl, Qt)
from PySide6.QtGui import (QBrush, QColor, QConicalGradient, QCursor,
    QFont, QFontDatabase, QGradient, QIcon,
    QImage, QKeySequence, QLinearGradient, QPainter,
    QPalette, QPixmap, QRadialGradient, QTransform)
from PySide6.QtSql import QSqlDatabase
from PySide6.QtWidgets import (QApplication, QHeaderView, QLabel, QLineEdit,
                               QPushButton, QSizePolicy, QTableWidget, QTableWidgetItem,
                               QWidget, QMessageBox)
from openpyxl.reader.excel import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from sqlalchemy import create_engine, text

from settings import DB_PATH, SAVE_DIR, EXCEL_TEMPLATE_PATH, MAIN_REPORT_PAGE
from time import time, sleep
from openpyxl.styles import Border, Side

class Ui_Form(object):
    def create_report(self):
        print("SHIT STARTED")

        db_connection = create_engine(f'sqlite:///{DB_PATH}').connect()
        df = pd.read_sql(text(f'SELECT * FROM papers'), db_connection).astype(str)



        # ----------------------------------------------------------------------------------дефолтные прелбразования данныхз ради некривого форматирования--------------------------------------------------------------------------------------------------------------------------

        # ----------------------------------------------------------------------------------создаем диреткории если нет и копируем шаблон, затем заполняем его--------------------------------------------------------------------------------------------------------------------------
        os.makedirs(SAVE_DIR, exist_ok=True)
        filename = os.path.join(SAVE_DIR, f'Платёжный чернослав{int(time())}.xlsx')
        shutil.copy(EXCEL_TEMPLATE_PATH, filename)
        current_report = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

        # заполнение данными
        # with open(NAMES_TXT_PATH, 'r', encoding='utf-8') as names_file:
        #     names = [line.strip() for line in names_file.readlines()]
        #self.employee_data = settings.get_employee_data(self.worker_pk)
        #self.employee_data = self.employee_data.replace("  ", "", regex=True)
        names = {

            'DATE':str(datetime.date.today().isoformat())


        }

        page = current_report.sheets[MAIN_REPORT_PAGE]
        print(df.columns)
        df_widght=len(df.columns)
        df_height =len(df)

        thin = Side(border_style="thin", color="000000")
        thick = Side(border_style="thick", color="000000")

        for i in range(df_height):
            for j in range(df_widght):
                page.cell(i+df_widght,j+df_height).border=Border(left=thick, right=thick, top=thick, bottom=thick)
                page.cell(i+df_widght,j+df_height).value=df.iloc[i][j]


        message = QMessageBox()
        message.setText("loh")
        message.show()
        sleep(1.5)

        def fill_placeholders(names, cell):
            if isinstance(cell.value, str):
                for placeholder, name in names.items():
                    cell.value = cell.value.replace(placeholder, name)

        for row in range(1, page.max_row + 1):
            for col in range(1, page.max_column + 1):
                fill_placeholders(names, page.cell(row, col))



        current_report.close()



    def fill_papers_table(self):
        TABLE_ROW_LIMIT = 10
        db_connection = create_engine(f'sqlite:///{DB_PATH}').connect()

        data_for_table = pd.read_sql(text(f'SELECT * FROM papers'), db_connection).astype(str)
        self.tableWidget.setRowCount(len(data_for_table))
        self.tableWidget.setColumnCount(len(data_for_table.columns))
        print(data_for_table,len(data_for_table),(len(data_for_table.columns)))
        for col_num in range(len(data_for_table.columns)):
            for row_num in range(len(data_for_table)):
                self.tableWidget.setItem(row_num, col_num,
                                                 QTableWidgetItem(data_for_table.iloc[row_num, col_num]))
    def setupUi(self, Form):
        if not Form.objectName():
            Form.setObjectName(u"Form")
        Form.resize(724, 507)
        self.addPushButton = QPushButton(Form)
        self.addPushButton.setObjectName(u"addPushButton")
        self.addPushButton.setGeometry(QRect(40, 420, 111, 61))
        self.tableWidget = QTableWidget(Form)
        self.tableWidget.setObjectName(u"tableWidget")
        self.tableWidget.setGeometry(QRect(30, 110, 651, 291))
        self.editPushButton = QPushButton(Form)
        self.editPushButton.setObjectName(u"editPushButton")
        self.editPushButton.setGeometry(QRect(180, 420, 111, 61))
        self.deletePushButton = QPushButton(Form)
        self.deletePushButton.setObjectName(u"deletePushButton")
        self.deletePushButton.setGeometry(QRect(580, 420, 111, 61))
        self.searchPushButton = QPushButton(Form)
        self.searchPushButton.setObjectName(u"searchPushButton")
        self.searchPushButton.setGeometry(QRect(310, 60, 111, 31))
        self.label = QLabel(Form)
        self.label.setObjectName(u"label")
        self.label.setGeometry(QRect(40, 20, 281, 31))
        font = QFont()
        font.setPointSize(19)
        self.label.setFont(font)
        self.searchLineEdit = QLineEdit(Form)
        self.searchLineEdit.setObjectName(u"searchLineEdit")
        self.searchLineEdit.setGeometry(QRect(40, 70, 251, 21))

        self.reportPushButton = QPushButton(Form, clicked = lambda:self.create_report() )
        self.reportPushButton.setObjectName(u"reportPushButton")
        self.reportPushButton.setGeometry(QRect(310, 420, 111, 61))

        self.fill_papers_table()

        self.retranslateUi(Form)

        QMetaObject.connectSlotsByName(Form)
    # setupUi

    def retranslateUi(self, Form):
        Form.setWindowTitle(QCoreApplication.translate("Form", u"Form", None))
        self.addPushButton.setText(QCoreApplication.translate("Form", u"\u0414\u043e\u0431\u0430\u0432\u0438\u0442\u044c", None))
        self.editPushButton.setText(QCoreApplication.translate("Form", u"\u0440\u0435\u0434\u0430\u043a\u0442\u0438\u0440\u043e\u0432\u0430\u0442\u044c", None))
        self.deletePushButton.setText(QCoreApplication.translate("Form", u"\u0423\u0434\u0430\u043b\u0438\u0442\u044c", None))
        self.searchPushButton.setText(QCoreApplication.translate("Form", u"\u0438\u0441\u043a\u0430\u0442\u044c", None))
        self.label.setText(QCoreApplication.translate("Form", u"\u0421\u043f\u0440\u0430\u0432\u043e\u0447\u043d\u0438\u043a \u0411\u0443\u043c\u0430\u0433", None))
    # retranslateUi

